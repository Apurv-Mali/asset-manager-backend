from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.response import Response
from rest_framework import status, viewsets, generics
from rest_framework.viewsets import ModelViewSet
from .models import Asset, AssetManager, Manager, TripDetails, DieselEntry, Breakdown, Mechanic, DieselStock, Operator, MonthlyRent
from .serializers import AssetSerializer, AssetManagerSerializer, ManagerSerializer, MechanicSerializer, SimpleAssetSerializer, ManagerWithAssetsSerializer, TripDetailsSerializer, ViewTripDetailsSerializer, DieselEntrySerializer, BreakdownSerializer, BreakdownReportSerializer, DieselReportSerializer, DieselStockSerializer, OperatorSerializer, MonthlyRentSerializer, TipperMonthlyReportSerializer, ExcavatorMonthlyReportSerializer, OtherMonthlyReportSerializer
from rest_framework.permissions import IsAuthenticated
from .permissions import IsManager
from .pagination import DefaultPagination
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.views import APIView
from rest_framework.exceptions import PermissionDenied
from django.db.models import ExpressionWrapper
from decimal import Decimal, ROUND_HALF_UP
from django.db.models import F
from rest_framework import status
from datetime import datetime, timedelta, time
from django.db.models import Sum, F, Value, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from openpyxl import Workbook
import io
from django.utils.timezone import make_aware
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.http import HttpResponse
from collections import defaultdict



class AssetViewSet(ModelViewSet):
    serializer_class = AssetSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    pagination_class = DefaultPagination
    search_fields = ['name','type']
    ordering_fields = ['name','type']

    def get_queryset(self):
        user = self.request.user

        if user.is_superuser:
            return Asset.objects.all()

        if hasattr(user, 'manager'):
            return Asset.objects.filter(asset_assignments__manager=user.manager)
        
        if hasattr(user, 'mechanic'):
            return Asset.objects.filter(breakdown_occurred__isnull=False).distinct()

        return Asset.objects.none()

    def create(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            return Response({"detail": "Not authorized to create assets."}, status=403)
        return super().create(request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            return Response({"detail": "Not authorized to delete assets."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            return Response({"detail": "Not authorized to update assets."}, status=status.HTTP_403_FORBIDDEN)

        allowed_fields = ['insurance_validity', 'monthly_rent', 'fitness_validity']
        data = {key: value for key, value in request.data.items() if key in allowed_fields}

        if not data:
            return Response({"detail": "No valid fields to update."}, status=status.HTTP_400_BAD_REQUEST)

        return super().partial_update(request, *args, **kwargs, partial=True)
    
class ManagerViewSet(ModelViewSet):
    queryset = Manager.objects.all()
    serializer_class = ManagerSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    pagination_class = DefaultPagination
    search_fields = ['name']

class MechanicViewSet(ModelViewSet):
    queryset = Mechanic.objects.all()
    serializer_class = MechanicSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    pagination_class = DefaultPagination
    search_fields = ['name']

class OperatorViewSet(ModelViewSet):
    queryset = Operator.objects.all()
    serializer_class = OperatorSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    pagination_class = DefaultPagination
    search_fields = ['name','type']

class MonthlyRentViewSet(ModelViewSet):
    queryset = MonthlyRent.objects.all()
    serializer_class = MonthlyRentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    pagination_class = DefaultPagination

class AssetManagerViewSet(viewsets.ModelViewSet):
    queryset = AssetManager.objects.all()
    serializer_class = AssetManagerSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
class UnassignedAssetsListView(ModelViewSet):
    queryset = Asset.objects.filter(asset_assignments__isnull=True)
    serializer_class = SimpleAssetSerializer
    permission_classes = [IsAuthenticated]


class RevokeAssetsViewSet(ModelViewSet):
    queryset = Manager.objects.all()
    serializer_class = ManagerWithAssetsSerializer
    permission_classes = [IsAuthenticated]


class TripDetailViewSet(ModelViewSet):
    serializer_class = TripDetailsSerializer
    queryset = TripDetails.objects.all()
    permission_classes = [IsManager]
    

    def get_serializer_class(self):
        if(self.request.method=='POST'):
            return TripDetailsSerializer
        if(self.request.method == 'GET'):
            return ViewTripDetailsSerializer
        return TripDetailsSerializer
    
    def get_queryset(self):
        queryset = TripDetails.objects.filter(manager=self.request.user.manager)
        return queryset.order_by('-date')[:5]
    
    def retrieve(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        instance = TripDetails.objects.get(pk=pk)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)
    
    def destroy(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        try:
            trip = TripDetails.objects.get(pk=pk)
            if trip.manager != request.user.manager:
                return Response({"error": "Unauthorized deletion attempt."}, status=status.HTTP_403_FORBIDDEN)
            trip.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except TripDetails.DoesNotExist:
            return Response({"error": "Trip not found."}, status=status.HTTP_404_NOT_FOUND)
        
    def update(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        try:
            trip = TripDetails.objects.get(pk=pk)

            if trip.manager != request.user.manager:
                return Response({"error": "Unauthorized update attempt."}, status=status.HTTP_403_FORBIDDEN)

            serializer = self.get_serializer(trip, data=request.data, partial=True)  # Allow partial updates

            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except TripDetails.DoesNotExist:
            return Response({"error": "Trip not found."}, status=status.HTTP_404_NOT_FOUND)
        
class ReceivableTripViewSet(ModelViewSet):
    queryset = TripDetails.objects.all()
    serializer_class = ViewTripDetailsSerializer
    pagination_class = DefaultPagination
    permission_classes = [IsManager]

    def get_queryset(self):
        queryset = TripDetails.objects.filter(receiver=self.request.user.manager)
        return queryset.order_by('-date')[:5]

class LoggedInManagerView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        if user.role == 'manager':
            try:
                manager = Manager.objects.get(user=user)
                return Response({'id': manager.id})
            except Manager.DoesNotExist:
                return Response({'detail': 'Manager profile not found.'}, status=404)
        else:
            return Response({'detail': 'You are not authorized as a manager.'}, status=403)
        
class LoggedInMechanicView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        if user.role == 'mechanic':
            try:
                mechanic = Mechanic.objects.get(user=user)
                return Response({'id': mechanic.id})
            except Mechanic.DoesNotExist:
                return Response({'detail': 'Mechanic profile not found.'}, status=404)
        else:
            return Response({'detail': 'You are not authorized as a mechanic.'}, status=403)
        
class DieselEntryViewSet(ModelViewSet):
    queryset = DieselEntry.objects.all()
    serializer_class = DieselEntrySerializer
    permission_classes = [IsManager]

    def get_queryset(self):
        queryset = DieselEntry.objects.filter(manager=self.request.user.manager)
        return queryset.order_by('-date')[:5]
    
    def destroy(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        try:
            Diesel = DieselEntry.objects.get(pk=pk)
            if Diesel.manager != request.user.manager:
                return Response({"error": "Unauthorized deletion attempt."}, status=status.HTTP_403_FORBIDDEN)
            Diesel.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Diesel.DoesNotExist:
            return Response({"error": "Diesel Entry not found."}, status=status.HTTP_404_NOT_FOUND)
        
    def retrieve(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        instance = DieselEntry.objects.get(pk=pk)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)
    
    def update(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        try:
            Diesel = DieselEntry.objects.get(pk=pk)

            # Check if the trip belongs to the current manager
            if Diesel.manager != request.user.manager:
                return Response({"error": "Unauthorized update attempt."}, status=status.HTTP_403_FORBIDDEN)

            # Use the serializer to validate and update the trip
            # Set partial=True to allow partial updates
            serializer = self.get_serializer(Diesel, data=request.data, partial=True)  # Allow partial updates

            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except DieselEntry.DoesNotExist:
            return Response({"error": "Diesel Entry not found."}, status=status.HTTP_404_NOT_FOUND)
        
class BreakdownViewSet(ModelViewSet):
    queryset = Breakdown.objects.filter(status__in=['Pending', 'Partially Completed'])
    serializer_class = BreakdownSerializer
    permission_classes = [IsAuthenticated]

class BreakdownReportViewSet(ModelViewSet):
    http_method_names = ['get']
    serializer_class = BreakdownReportSerializer
    pagination_class = DefaultPagination
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['asset__name','site','status']

    def get_queryset(self):
        user = self.request.user

        if user.is_superuser:
            return Breakdown.objects.all()

        if hasattr(user, 'manager'):
            return Breakdown.objects.filter(manager=user.manager)
        
        if hasattr(user, 'mechanic'):
            return Breakdown.objects.filter(mechanic=user.mechanic)

        return Breakdown.objects.none()
        
class DieselReportViewSet(ModelViewSet):
    http_method_names = ['get']
    queryset = DieselEntry.objects.all()
    permission_classes = [IsAuthenticated]
    serializer_class = DieselReportSerializer
    pagination_class = DefaultPagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['asset__registration_no',]
    ordering_fields = ['id']

class DieselStockViewSet(ModelViewSet):
    queryset = DieselStock.objects.all()
    serializer_class = DieselStockSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    pagination_class = DefaultPagination
    search_fields = ['challan_no']

def round_decimal(value):
    return round(value, 2)


class TipperMonthlyReportView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        month = request.query_params.get("month")  # expect "YYYY-MM"
        if not month:
            return Response({"error": "Month is required (format YYYY-MM)."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = datetime.strptime(month + "-01", "%Y-%m-%d").date()
            next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
            end_date = next_month - timedelta(days=1)
            start_datetime = make_aware(datetime.combine(start_date, time.min))
            end_datetime = make_aware(datetime.combine(end_date, time.max))
        except ValueError:
            return Response({"error": "Invalid month format. Use YYYY-MM."}, status=status.HTTP_400_BAD_REQUEST)

        tipper_assets = Asset.objects.filter(type__iexact="Tipper")
        report_rows = []

        for asset in tipper_assets:
            # Get manager
            latest_assign = (
                AssetManager.objects.filter(asset=asset, date_assigned__lte=end_datetime)
                .select_related("manager")
                .order_by("-date_assigned")
                .first()
            )
            manager_name = str(latest_assign.manager) if latest_assign else "Unassigned"

            # Get trips and diesel entries
            trips = TripDetails.objects.filter(asset=asset, date__range=(start_datetime, end_datetime))
            diesel_entries = DieselEntry.objects.filter(asset=asset, date__range=(start_datetime, end_datetime))

            # Group trips by date → then by material
            trips_by_day_material = defaultdict(lambda: defaultdict(list))
            for trip in trips:
                trip_date = trip.date.date()
                trips_by_day_material[trip_date][trip.material].append(trip)

            # Diesel data grouped by day
            diesel_by_day = defaultdict(list)
            for de in diesel_entries:
                diesel_by_day[de.date.date()].append(de)

            # Aggregate per material
            material_summary = defaultdict(lambda: {
                "material_quantity": Decimal("0"),
                "rate": Decimal("0"),
                "diesel_consumed": Decimal("0"),
                "diesel_cost": Decimal("0"),
                "distance": Decimal("0"),
            })

            for trip_date, material_dict in trips_by_day_material.items():
                total_distance_day = 0
                distance_by_material = {}

                for material, trips_list in material_dict.items():
                    dist = sum(Decimal(trip.distance or 0) for trip in trips_list)
                    distance_by_material[material] = dist
                    total_distance_day += dist

                diesel_liters = sum((d.quantity or 0) for d in diesel_by_day.get(trip_date, []))
                diesel_cost = sum((d.quantity * d.rate) for d in diesel_by_day.get(trip_date, []))

                only_one_material = len(material_dict) == 1

                for material, trips_list in material_dict.items():
                    qty = sum((trip.net_weight or 0) for trip in trips_list)
                    rate = trips_list[0].rate or Decimal("0")
                    dist = distance_by_material[material]

                    if only_one_material:
                        mat_diesel = diesel_liters
                        mat_diesel_cost = diesel_cost
                    elif total_distance_day > 0:
                        proportion = dist / total_distance_day
                        mat_diesel = diesel_liters * proportion
                        mat_diesel_cost = diesel_cost * proportion
                    else:
                        mat_diesel = Decimal("0")
                        mat_diesel_cost = Decimal("0")

                    summary = material_summary[material]
                    summary["material_quantity"] += qty
                    summary["rate"] = rate  # Assume latest or first rate
                    summary["diesel_consumed"] += mat_diesel
                    summary["diesel_cost"] += mat_diesel_cost
                    summary["distance"] += dist

            # Append all rows
            for material, data in material_summary.items():
                amount = data["material_quantity"] * data["rate"]
                final_amount = amount + data["diesel_cost"]

                report_rows.append({
                    "asset_name": f"{asset.name} - {asset.registration_no}",
                    "manager": manager_name,
                    "material": material,
                    "material_quantity": data["material_quantity"],
                    "rate": data["rate"],
                    "diesel_consumed": round_decimal(data["diesel_consumed"]),
                    "diesel_cost": round_decimal(data["diesel_cost"]),
                    "distance": data["distance"],
                    "amount": round_decimal(amount),
                    "final_amount": round_decimal(final_amount),
                    "status": "Active" if trips else "Idle",
                })

        serializer = TipperMonthlyReportSerializer(data=report_rows, many=True)
        serializer.is_valid(raise_exception=True)
        return Response(serializer.data)


class TipperMonthlyReportExportView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        # Get the processed report data from the main view
        data_response = TipperMonthlyReportView().get(request)
        data = data_response.data

        # Create a new Excel workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tipper Monthly Report"

        # Define headers (without breakdowns)
        headers = [
            "Asset Name", "Manager", "Material", "Material Quantity", "Rate",
            "Diesel Consumed", "Diesel Cost", "Distance", "Amount", "Final Amount", "Status"
        ]
        ws.append(headers)

        # Add rows to sheet
        for row in data:
            ws.append([
                row["asset_name"],
                row["manager"],
                row["material"],
                float(row["material_quantity"]),
                float(row["rate"]),
                float(row["diesel_consumed"]),
                float(row["diesel_cost"]),
                float(row["distance"]),
                float(row["amount"]),
                float(row["final_amount"]),
                row["status"],
            ])

        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        # Prepare file response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        month = request.query_params.get("month", "report")
        response["Content-Disposition"] = f'attachment; filename="tipper_report_{month}.xlsx"'
        wb.save(response)

        return response
    
class ExcavatorMonthlyReportView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        month = request.query_params.get("month")  # Expected "YYYY-MM"
        if not month:
            return Response(
                {"error": "Month is required (format YYYY-MM)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            start_date = datetime.strptime(month + "-01", "%Y-%m-%d").date()
            next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
            end_date = next_month - timedelta(days=1)
            start_datetime = make_aware(datetime.combine(start_date, time.min))
            end_datetime = make_aware(datetime.combine(end_date, time.max))
        except ValueError:
            return Response(
                {"error": "Invalid month format. Use YYYY-MM."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        excavator_assets = Asset.objects.filter(type__iexact="Excavator")
        report_rows = []

        for asset in excavator_assets:
            latest_assign = (
                AssetManager.objects.filter(asset=asset, date_assigned__lte=end_datetime)
                .select_related("manager")
                .order_by("-date_assigned")
                .first()
            )
            manager_name = str(latest_assign.manager) if latest_assign else "Unassigned"

            trips_qs = TripDetails.objects.filter(asset=asset, date__range=(start_datetime, end_datetime))
            
            if not trips_qs.exists():
                continue  # 🔁 Skip this asset if no trips in month

            diesel_qs = DieselEntry.objects.filter(asset=asset, date__range=(start_datetime, end_datetime))

            total_hours = (
                trips_qs.aggregate(
                    hours=Sum("hours", output_field=DecimalField())
                )["hours"] or Decimal("0")
            )
            total_shifts = (
                trips_qs.aggregate(
                    shifts=Sum("shift", output_field=DecimalField())
                )["shifts"] or Decimal("0")
            )

            diesel_qty = (
                diesel_qs.aggregate(
                    total=Sum("quantity", output_field=DecimalField())
                )["total"] or Decimal("0")
            )
            diesel_amt = (
                diesel_qs.aggregate(
                    cost=Sum(
                        ExpressionWrapper(F("quantity") * F("rate"), output_field=DecimalField())
                    )
                )["cost"] or Decimal("0")
            )

            monthly_charge = asset.rate_per_month or Decimal("0")
            shift_charge = asset.rate_per_shift or Decimal("0")

            amount = total_shifts * shift_charge
            final_amount = amount + diesel_amt

            report_rows.append({
                "asset_name": f"{asset.name} - {asset.registration_no}",
                "manager": manager_name,
                "working_hours": round_decimal(total_hours),
                "monthly_charge": round_decimal(monthly_charge),
                "shift_charge": round_decimal(shift_charge),
                "no_of_shifts": round_decimal(total_shifts),
                "amount": round_decimal(amount),
                "diesel_quantity": round_decimal(diesel_qty),
                "diesel_amount": round_decimal(diesel_amt),
                "final_amount": round_decimal(final_amount),
            })

        serializer = ExcavatorMonthlyReportSerializer(data=report_rows, many=True)
        serializer.is_valid(raise_exception=True)
        return Response(serializer.data)
    
class ExcavatorMonthlyReportExportView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data_response = ExcavatorMonthlyReportView().get(request)
        data = data_response.data

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Excavator Monthly Report"

        # Column headers
        headers = [
            "Asset Name", "Manager", "Working Hours", "Monthly Charge", "Shift Charge",
            "No. of Shifts", "Amount", "Diesel Quantity", "Diesel Amount", "Final Amount"
        ]
        ws.append(headers)

        # Fill rows
        for row in data:
            ws.append([
                row["asset_name"],
                row["manager"],
                float(row["working_hours"]),
                float(row["monthly_charge"]),
                float(row["shift_charge"]),
                float(row["no_of_shifts"]),
                float(row["amount"]),
                float(row["diesel_quantity"]),
                float(row["diesel_amount"]),
                float(row["final_amount"]),
            ])

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        # Prepare response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        month = request.query_params.get("month", "report")
        response["Content-Disposition"] = f'attachment; filename="excavator_report_{month}.xlsx"'
        wb.save(response)

        return response
    
class OtherAssetsMonthlyReportView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        month = request.query_params.get("month")  # Expected "YYYY-MM"
        if not month:
            return Response(
                {"error": "Month is required (format YYYY-MM)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            start_date = datetime.strptime(month + "-01", "%Y-%m-%d").date()
            next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
            end_date = next_month - timedelta(days=1)
            start_datetime = make_aware(datetime.combine(start_date, time.min))
            end_datetime = make_aware(datetime.combine(end_date, time.max))
        except ValueError:
            return Response(
                {"error": "Invalid month format. Use YYYY-MM."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        other_assets = Asset.objects.exclude(type__iexact="Excavator").exclude(type__iexact="Tipper")
        report_rows = []

        for asset in other_assets:
            latest_assign = (
                AssetManager.objects.filter(asset=asset, date_assigned__lte=end_datetime)
                .select_related("manager")
                .order_by("-date_assigned")
                .first()
            )
            manager_name = str(latest_assign.manager) if latest_assign else "Unassigned"

            trips_qs = TripDetails.objects.filter(asset=asset, date__range=(start_datetime, end_datetime))

            if not trips_qs.exists():
                continue  # ⛔ Skip asset if no trips in this month

            diesel_qs = DieselEntry.objects.filter(asset=asset, date__range=(start_datetime, end_datetime))

            total_shifts = trips_qs.aggregate(
                shifts=Sum("shift", output_field=DecimalField())
            )["shifts"] or Decimal("0")

            diesel_qty = diesel_qs.aggregate(
                total=Sum("quantity", output_field=DecimalField())
            )["total"] or Decimal("0")

            diesel_amt = diesel_qs.aggregate(
                cost=Sum(
                    ExpressionWrapper(F("quantity") * F("rate"), output_field=DecimalField())
                )
            )["cost"] or Decimal("0")

            monthly_charge = asset.rate_per_month or Decimal("0")
            shift_charge = asset.rate_per_shift or Decimal("0")

            amount = total_shifts * shift_charge
            final_amount = amount + diesel_amt

            report_rows.append({
                "asset_name": f"{asset.name} - {asset.registration_no}",
                "manager": manager_name,
                "monthly_charge": round_decimal(monthly_charge),
                "shift_charge": round_decimal(shift_charge),
                "no_of_shifts": round_decimal(total_shifts),
                "amount": round_decimal(amount),
                "diesel_quantity": round_decimal(diesel_qty),
                "diesel_amount": round_decimal(diesel_amt),
                "final_amount": round_decimal(final_amount),
            })

        serializer = OtherMonthlyReportSerializer(data=report_rows, many=True)
        serializer.is_valid(raise_exception=True)
        return Response(serializer.data)

class OtherAssetsMonthlyReportExportView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data_response = OtherAssetsMonthlyReportView().get(request)
        data = data_response.data

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Other Assets Monthly Report"

        # Define column headers
        headers = [
            "Asset Name", "Manager", "Monthly Charge", "Shift Charge",
            "No. of Shifts", "Amount", "Diesel Quantity",
            "Diesel Amount", "Final Amount"
        ]
        ws.append(headers)

        # Fill rows with data
        for row in data:
            ws.append([
                row["asset_name"],
                row["manager"],
                float(row["monthly_charge"]),
                float(row["shift_charge"]),
                float(row["no_of_shifts"]),
                float(row["amount"]),
                float(row["diesel_quantity"]),
                float(row["diesel_amount"]),
                float(row["final_amount"]),
            ])

        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        # Prepare response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        month = request.query_params.get("month", "report")
        response["Content-Disposition"] = f'attachment; filename="other_assets_report_{month}.xlsx"'
        wb.save(response)

        return response
    
class CompleteAssetMonthlyReportView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        month = request.query_params.get("month")  # Expected "YYYY-MM"
        if not month:
            return Response({"error": "Month is required (format YYYY-MM)."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = datetime.strptime(month + "-01", "%Y-%m-%d").date()
            next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
            end_date = next_month - timedelta(days=1)
            start_datetime = make_aware(datetime.combine(start_date, time.min))
            end_datetime = make_aware(datetime.combine(end_date, time.max))
        except ValueError:
            return Response({"error": "Invalid month format. Use YYYY-MM."}, status=status.HTTP_400_BAD_REQUEST)

        report = []

        for asset in Asset.objects.all():
            trips = TripDetails.objects.filter(asset=asset, date__range=(start_datetime, end_datetime))
            diesel_entries = DieselEntry.objects.filter(asset=asset, date__range=(start_datetime, end_datetime))
            breakdowns = asset.breakdown_occurred.filter(date__range=(start_datetime, end_datetime))

            # Check asset usage
            status_str = "Active" if trips.exists() else "Idle"

            # Total running
            if trips.filter(distance__isnull=False).exists():
                running_value = trips.aggregate(total=Sum("distance", output_field=DecimalField()))["total"] or Decimal("0")
                running_str = f"{round_decimal(running_value)} kms"
            elif trips.filter(hours__isnull=False).exists():
                running_value = trips.aggregate(total=Sum("hours", output_field=DecimalField()))["total"] or Decimal("0")
                running_str = f"{round_decimal(running_value)} hrs"
            elif trips.filter(shift__isnull=False).exists():
                running_value = trips.aggregate(total=Sum("shift", output_field=DecimalField()))["total"] or Decimal("0")
                running_str = f"{round_decimal(running_value)} shifts"
            else:
                running_str = "0"

            # Fixed revenue
            fixed_revenue = asset.rate_per_month or Decimal("0")

            # Transport revenue
            if trips.filter(net_weight__isnull=False).exists():
                transport_revenue = trips.aggregate(
                    total=Sum(ExpressionWrapper(F("rate") * F("net_weight"), output_field=DecimalField()))
                )["total"] or Decimal("0"
                )
            elif trips.filter(hours__isnull=False).exists():
                transport_revenue = trips.aggregate(
                    total=Sum(ExpressionWrapper(F("rate") * F("hours"), output_field=DecimalField()))
                )["total"] or Decimal("0"
                )
            else:
                transport_revenue = trips.aggregate(
                    total=Sum(ExpressionWrapper(F("rate") * F("shift"), output_field=DecimalField()))
                )["total"] or Decimal("0"
                )

            # Diesel quantity & amount
            diesel_quantity = diesel_entries.aggregate(total=Sum("quantity", output_field=DecimalField()))["total"] or Decimal("0")
            diesel_amount = diesel_entries.aggregate(
                total=Sum(ExpressionWrapper(F("quantity") * F("rate"), output_field=DecimalField()))
            )["total"] or Decimal("0")

            # Maintenance = cost + manpower_cost
            maintenance_cost = breakdowns.aggregate(
                total=Sum(ExpressionWrapper(
                    (F("cost") + F("manpower_cost")),
                    output_field=DecimalField()
                ))
            )["total"] or Decimal("0")

            report.append({
                "asset": f"{asset.name} - {asset.registration_no}",
                "total_running": running_str,
                "fixed_revenue": round_decimal(fixed_revenue),
                "revenue_transport": round_decimal(transport_revenue),
                "diesel_quantity": round_decimal(diesel_quantity),
                "diesel_amount": round_decimal(diesel_amount),
                "maintenance": round_decimal(maintenance_cost),
                "insurance": round_decimal(asset.insurance_amount),
                "tax": round_decimal(asset.road_tax_amount),
                "fitness": round_decimal(asset.fitness_amount),
                "permit": round_decimal(asset.permit_amount),
                "emi": round_decimal(asset.emi_amount),
                "status": status_str,
            })

        return Response(report)

class CompleteAssetMonthlyReportExcelView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        month = request.query_params.get("month")  # e.g., "2025-06"
        if not month:
            return Response({"error": "Month is required in YYYY-MM format."}, status=400)

        try:
            start_date = datetime.strptime(month + "-01", "%Y-%m-%d").date()
            next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
            end_date = next_month - timedelta(days=1)
            start_datetime = make_aware(datetime.combine(start_date, time.min))
            end_datetime = make_aware(datetime.combine(end_date, time.max))
        except ValueError:
            return Response({"error": "Invalid month format. Use YYYY-MM."}, status=400)

        from .models import Asset, TripDetails, DieselEntry, Breakdown  # adjust as per your structure

        assets = Asset.objects.all()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Asset Monthly Report"

        headers = [
            "Asset", "Total Running", "Fixed Revenue", "Revenue Transport",
            "Diesel Quantity", "Diesel Amount", "Maintenance", "Insurance",
            "Tax", "Fitness", "Permit", "EMI", "Status"
        ]

        # Add bold headers
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        row_num = 2

        for asset in assets:
            trips = TripDetails.objects.filter(asset=asset, date__range=(start_datetime, end_datetime))
            diesel = DieselEntry.objects.filter(asset=asset, date__range=(start_datetime, end_datetime))
            breakdowns = Breakdown.objects.filter(asset=asset, date__range=(start_datetime, end_datetime))

            # Total running
            total_distance = trips.aggregate(d=Sum("distance", output_field=DecimalField()))["d"]
            total_hours = trips.aggregate(h=Sum("hours", output_field=DecimalField()))["h"]
            total_shifts = trips.aggregate(s=Sum("shift", output_field=DecimalField()))["s"]
            if total_distance:
                running_value = f"{round(total_distance)} kms"
            elif total_hours:
                running_value = f"{round(total_hours)} hrs"
            elif total_shifts:
                running_value = f"{round(total_shifts)} shifts"
            else:
                continue  # Skip if no trips

            # Fixed Revenue
            fixed_revenue = asset.rate_per_month or Decimal("0")

            # Revenue Transport
            if total_distance:
                revenue = trips.aggregate(
                    revenue=Sum(ExpressionWrapper(F("rate") * F("net_weight"), output_field=DecimalField()))
                )["revenue"] or Decimal("0")
            elif total_hours:
                revenue = trips.aggregate(
                    revenue=Sum(ExpressionWrapper(F("rate") * F("hours"), output_field=DecimalField()))
                )["revenue"] or Decimal("0")
            elif total_shifts:
                revenue = trips.aggregate(
                    revenue=Sum(ExpressionWrapper(F("rate") * F("shift"), output_field=DecimalField()))
                )["revenue"] or Decimal("0")
            else:
                revenue = Decimal("0")

            # Diesel
            diesel_qty = diesel.aggregate(q=Sum("quantity", output_field=DecimalField()))["q"] or Decimal("0")
            diesel_amt = diesel.aggregate(
                a=Sum(ExpressionWrapper(F("quantity") * F("rate"), output_field=DecimalField()))
            )["a"] or Decimal("0")

            # Maintenance
            breakdown_cost = breakdowns.aggregate(
                total=Sum(
                    ExpressionWrapper(
                        Coalesce(F("cost"), 0) + Coalesce(F("manpower_cost"), 0),
                        output_field=DecimalField()
                    )
                )
            )["total"] or Decimal("0")

            # Additional Charges
            insurance = asset.insurance_amount or Decimal("0")
            tax = asset.road_tax_amount or Decimal("0")
            fitness = asset.fitness_amount or Decimal("0")
            permit = asset.permit_amount or Decimal("0")
            emi = asset.emi_amount or Decimal("0")

            # Status
            status = "Active" if trips.exists() else "Idle"

            values = [
                f"{asset.name} - {asset.registration_no}",
                running_value,
                float(fixed_revenue),
                float(revenue),
                float(diesel_qty),
                float(diesel_amt),
                float(breakdown_cost),
                float(insurance),
                float(tax),
                float(fitness),
                float(permit),
                float(emi),
                status
            ]

            for col, val in enumerate(values, 1):
                sheet.cell(row=row_num, column=col).value = val

            row_num += 1

        # Auto-adjust column widths
        for col in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        # Prepare response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        filename = f"asset_monthly_report_{month}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response
    
from django.utils.timezone import localtime

class TripDetailsReportView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        month_param = request.query_params.get('month')
        asset_type = request.query_params.get('asset_type')  # Optional

        if not month_param:
            return Response({"error": "Month parameter is required in YYYY-MM format."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = datetime.strptime(month_param, "%Y-%m")
            # Compute the end of the month
            if start_date.month == 12:
                end_date = datetime(start_date.year + 1, 1, 1)
            else:
                end_date = datetime(start_date.year, start_date.month + 1, 1)

            start_date = make_aware(start_date)
            end_date = make_aware(end_date)

        except ValueError:
            return Response({"error": "Invalid month format. Use YYYY-MM."}, status=status.HTTP_400_BAD_REQUEST)

        queryset = TripDetails.objects.filter(
            date__gte=start_date,
            date__lt=end_date
        ).select_related('asset', 'manager', 'receiver', 'operator')

        if asset_type:
            queryset = queryset.filter(asset__type=asset_type)

        tipper_types = ["Tipper", "Trailor"]
        excavator_types = ["Excavator", "Paver", "Dozer", "Loader", "Grader", "Roller", "Backhoe Loader", "Farana"]
        transporter_types = ["B-Tempo", "Bus", "Camper", "Service Van", "Transit Mixer", "Water Tanker"]

        data = []

        for trip in queryset:
            a_type = trip.asset.type
            trip_time = localtime(trip.date)  # ✅ Convert UTC to IST

            base = {
                "asset": f"{trip.asset.name} - {trip.asset.registration_no}",
                "deal_type": trip.deal_type,
                "manager": trip.manager.name if trip.manager else None,
                "receiver": trip.receiver.name if trip.receiver else None,
                "operator": trip.operator.name if trip.operator else None,
                "date": trip_time.strftime("%d-%m-%Y"),
                "time": trip_time.strftime("%I:%M %p"),
            }

            if a_type in tipper_types:
                amount = float(trip.rate or 0) * float(trip.net_weight or 0)
                base.update({
                    "from_location": trip.from_location,
                    "to_location": trip.to_location,
                    "material": trip.material,
                    "rate": f"{trip.rate} per ton",
                    "distance": f"{trip.distance} kms" if trip.distance else None,
                    "net_weight": f"{trip.net_weight} ton" if trip.net_weight else None,
                    "amount": round(amount, 2),
                })

            elif a_type in excavator_types:
                amount = float(trip.rate or 0) * float(trip.hours or 0)
                base.update({
                    "from_location": trip.from_location,
                    "rate": trip.rate,
                    "hours": trip.hours,
                    "start_time": trip.start_time.strftime("%I:%M %p") if trip.start_time else None,
                    "end_time": trip.end_time.strftime("%I:%M %p") if trip.end_time else None,
                    "shift": trip.shift,
                    "amount": round(amount, 2),
                })

            elif a_type in transporter_types:
                amount = float(trip.rate or 0) * float(trip.shift or 0)
                base.update({
                    "from_location": trip.from_location,
                    "to_location": trip.to_location,
                    "rate": trip.rate,
                    "shift": trip.shift,
                    "amount": round(amount, 2),
                })

            data.append(base)

        return Response(data, status=status.HTTP_200_OK)

    
class TripDetailsReportExcelView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        month_param = request.query_params.get('month')
        asset_type = request.query_params.get('asset_type')

        if not month_param:
            return Response({"error": "Month parameter is required in YYYY-MM format."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = datetime.strptime(month_param, "%Y-%m")
            end_date = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
            start_date = make_aware(start_date)
            end_date = make_aware(end_date)
        except ValueError:
            return Response({"error": "Invalid month format. Use YYYY-MM."}, status=status.HTTP_400_BAD_REQUEST)

        queryset = TripDetails.objects.filter(
            date__gte=start_date,
            date__lt=end_date
        ).select_related('asset', 'manager', 'receiver', 'operator')

        if asset_type:
            queryset = queryset.filter(asset__type=asset_type)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Trip Report"

        tipper_types = ["Tipper", "Trailor"]
        excavator_types = ["Excavator", "Paver", "Dozer", "Loader", "Grader", "Roller", "Backhoe Loader", "Farana"]
        transporter_types = ["B-Tempo", "Bus", "Camper", "Service Van", "Transit Mixer", "Water Tanker"]

        headers = [
            "Asset", "Deal Type", "Manager", "Receiver", "Operator", "Date", "Time"
        ]

        if asset_type in tipper_types:
            headers += ["From", "To", "Material", "Rate", "Distance", "Net Weight", "Amount"]
        elif asset_type in excavator_types:
            headers += ["From", "Rate", "Hours", "Start Time", "End Time", "Shift", "Amount"]
        elif asset_type in transporter_types:
            headers += ["From", "To", "Rate", "Shift", "Amount"]
        else:
            headers += ["From", "To", "Rate", "Amount"]

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        row_num = 2
        for trip in queryset:
            a_type = trip.asset.type
            ist_date = localtime(trip.date)  # ✅ IST conversion here

            base = [
                f"{trip.asset.name} - {trip.asset.registration_no}",
                trip.deal_type,
                trip.manager.name if trip.manager else "",
                trip.receiver.name if trip.receiver else "",
                trip.operator.name if trip.operator else "",
                ist_date.strftime("%d-%m-%Y"),    # ✅ IST Date
                ist_date.strftime("%I:%M %p"),     # ✅ IST Time
            ]

            if a_type in tipper_types:
                amount = float(trip.rate or 0) * float(trip.net_weight or 0)
                base += [
                    trip.from_location,
                    trip.to_location,
                    trip.material,
                    float(trip.rate) if trip.rate else None,
                    float(trip.distance) if trip.distance else None,
                    float(trip.net_weight) if trip.net_weight else None,
                    round(amount, 2)
                ]

            elif a_type in excavator_types:
                amount = float(trip.rate or 0) * float(trip.hours or 0)
                base += [
                    trip.from_location,
                    float(trip.rate) if trip.rate else None,
                    trip.hours,
                    trip.start_time.strftime("%I:%M %p") if trip.start_time else "",
                    trip.end_time.strftime("%I:%M %p") if trip.end_time else "",
                    trip.shift,
                    round(amount, 2)
                ]

            elif a_type in transporter_types:
                amount = float(trip.rate or 0) * float(trip.shift or 0)
                base += [
                    trip.from_location,
                    trip.to_location,
                    float(trip.rate) if trip.rate else None,
                    trip.shift,
                    round(amount, 2)
                ]
            else:
                amount = float(trip.rate or 0)
                base += [
                    trip.from_location,
                    trip.to_location,
                    float(trip.rate) if trip.rate else None,
                    round(amount, 2)
                ]

            for col_num, value in enumerate(base, 1):
                sheet.cell(row=row_num, column=col_num).value = value
            row_num += 1

        for col in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        type_str = asset_type.lower().replace(" ", "_") if asset_type else "all"
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename=trip_report_{type_str}_{month_param}.xlsx'
        workbook.save(response)
        return response